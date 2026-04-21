"""
Generate Excel run template(s) from baseline STIX file(s).

Usage:
    python generate_template.py                    # all .stix in baseline_models/
    python generate_template.py baseline_models/bergambacht.stix   # one file
    python generate_template.py baseline_models/bergambacht.stix baseline_models/eemdijk.stix

This creates an Excel file next to each STIX file:
    baseline_models/<name>_runs.xlsx

Sheets:
    runs        - One row per run: description, calc methods to run
    materials   - One row per (run, material): strength model + parameters
    su_tables   - Registry of SuTable references
    results     - Filled in automatically by run_model.py
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from source.stix_io import (
    read_stix,
    get_soils,
    get_calc_settings_map,
    get_soil_pop_map,
    get_soil_layers_map,
)


def extract_soil_params(soil: dict) -> dict:
    """Extract all strength parameters from a soil object."""
    model_below = soil.get("ShearStrengthModelTypeBelowPhreaticLevel", "")
    model_above = soil.get("ShearStrengthModelTypeAbovePhreaticLevel", "")

    row = {
        "material_code": soil.get("Code", ""),
        "active_model": model_below,
        "gamma_dry": soil.get("VolumetricWeightAbovePhreaticLevel", ""),
        "gamma_wet": soil.get("VolumetricWeightBelowPhreaticLevel", ""),
        # Su / SHANSEP
        "Su_S": "",
        "Su_m": "",
        # MohrCoulomb
        "MC_phi": "",
        "MC_c": "",
        "MC_psi": "",
        # SuTable
        "su_table_key": "",
    }

    if model_below in ("Su", "SuShearStrengthModel"):
        m = soil.get("SuShearStrengthModel", {})
        row["Su_S"] = m.get("ShearStrengthRatio", "")
        row["Su_m"] = m.get("StrengthIncreaseExponent", "")

    elif model_below == "SuTable":
        m = soil.get("SuTable", {})
        row["Su_m"] = m.get("StrengthIncreaseExponent", "")
        row["su_table_key"] = ""  # user fills in

    elif model_below in ("MohrCoulombAdvanced", "MohrCoulombClassic"):
        key = (
            "MohrCoulombAdvancedShearStrengthModel"
            if model_below == "MohrCoulombAdvanced"
            else "MohrCoulombClassicShearStrengthModel"
        )
        m = soil.get(key, {})
        row["MC_phi"] = m.get("FrictionAngle", "")
        row["MC_c"] = m.get("Cohesion", "")
        if model_below == "MohrCoulombAdvanced":
            row["MC_psi"] = m.get("Dilatancy", "")

    return row


def build_runs_sheet(calc_methods: list) -> pd.DataFrame:
    """Build the runs sheet with one baseline row."""
    row = {
        "run_id": "baseline",
        "description": "Original model, no modifications",
        "notes": "",
    }
    for method in calc_methods:
        row[f"run_{method}"] = True
    return pd.DataFrame([row])


def build_materials_sheet(soils: list, pop_map: dict, layers_map: dict) -> pd.DataFrame:
    """Build the materials sheet with baseline parameters for all soils."""
    # layers_map: {scenario_label -> {soil_code -> [layer_label, ...]}}
    scenario_labels = sorted(layers_map.keys())  # ['s1', 's2', ...]

    rows = []
    for soil in soils:
        row = extract_soil_params(soil)
        row["run_id"] = "baseline"

        # One column per scenario showing which layers this soil occupies
        for sc in scenario_labels:
            assigned = layers_map[sc].get(row["material_code"], [])
            row[f"layer_{sc}"] = ", ".join(assigned) if assigned else ""

        # Summarise POP into two columns: state point labels and values
        pops = pop_map.get(row["material_code"], [])
        if pops:
            row["Layers"] = ", ".join(label for label, _ in pops)
            row["POP"] = ", ".join(str(val) for _, val in pops)
        else:
            row["Layers"] = ""
            row["POP"] = ""

        rows.append(row)

    df = pd.DataFrame(rows)
    layer_cols = [f"layer_{sc}" for sc in scenario_labels]
    cols = [
        "run_id",
        "material_code",
        *layer_cols,
        "active_model",
        "gamma_dry",
        "gamma_wet",
        "Su_S",
        "Su_m",
        "MC_phi",
        "MC_c",
        "MC_psi",
        "su_table_key",
        "Layers",
        "POP",
    ]
    df = df[cols]
    return df


def build_su_tables_sheet() -> pd.DataFrame:
    """Build the SuTable registry sheet (empty, user fills in)."""
    return pd.DataFrame(columns=["su_table_key", "description", "file"])


def build_results_sheet(calc_methods: list) -> pd.DataFrame:
    """Build the results sheet (empty, filled by run_model.py)."""
    cols = ["run_id", "model_name", "timestamp"] + [f"FoS_{m}" for m in calc_methods]
    return pd.DataFrame(columns=cols)


def style_sheet(ws, header_color="1F4E79"):
    """Apply basic styling to a worksheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=header_color)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 4, 40
        )


METHOD_LABELS = {
    "BishopBruteForce": "bishop",
    "UpliftVanParticleSwarm": "upliftvan",
    "SpencerGenetic": "spencer",
}

SHEET_COLORS = {
    "runs": "1F4E79",
    "materials": "375623",
    "su_tables": "7B3F00",
    "results": "4A235A",
}


def generate_for_stix(stix_path: Path) -> None:
    print(f"\nReading: {stix_path.name}")
    data = read_stix(stix_path)

    soils = get_soils(data)
    pop_map = get_soil_pop_map(data)
    layers_map = get_soil_layers_map(data)
    calc_map = get_calc_settings_map(data)
    calc_methods = [METHOD_LABELS.get(k, k) for k in calc_map.keys()]

    print(f"  Found {len(soils)} soils")
    print(f"  Found calc methods: {calc_methods}")

    df_runs = build_runs_sheet(calc_methods)
    df_materials = build_materials_sheet(soils, pop_map, layers_map)
    df_su_tables = build_su_tables_sheet()
    df_results = build_results_sheet(calc_methods)

    output_path = stix_path.with_name(stix_path.stem + "_runs.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_runs.to_excel(writer, sheet_name="runs", index=False)
        df_materials.to_excel(writer, sheet_name="materials", index=False)
        df_su_tables.to_excel(writer, sheet_name="su_tables", index=False)
        df_results.to_excel(writer, sheet_name="results", index=False)

    wb = load_workbook(output_path)
    for sheet_name, color in SHEET_COLORS.items():
        if sheet_name in wb.sheetnames:
            style_sheet(wb[sheet_name], header_color=color)
    wb.save(output_path)

    print(f"  ✓ Template created: {output_path.name}")


def main():
    project_root = Path(__file__).parent

    if len(sys.argv) > 1:
        stix_paths = [project_root / arg for arg in sys.argv[1:]]
    else:
        stix_paths = sorted((project_root / "baseline_models").glob("*.stix"))

    if not stix_paths:
        print("ERROR: No STIX files found in baseline_models/")
        sys.exit(1)

    for stix_path in stix_paths:
        if not stix_path.exists():
            print(f"  WARNING: not found, skipping: {stix_path}")
            continue
        generate_for_stix(stix_path)

    print(f"\nDone. {len(stix_paths)} template(s) generated.")
    print("  runs      - Define your runs here (add rows for run_1, run_2, ...)")
    print("  materials - Define parameter changes per run (empty cell = use baseline)")
    print("  su_tables - Register SuTable JSON files here")
    print("  results   - Filled automatically by run_model.py")


if __name__ == "__main__":
    main()
